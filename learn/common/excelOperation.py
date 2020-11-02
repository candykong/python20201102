#!/usr/bin/python
# -*- coding:UTF-8 -*-

import requests
import json
import xlrd
import xlwt
import openpyxl
from openpyxl import load_workbook


#读取excel-按文件-sheet名-行和列读取数据
def outputExcel(path,sheet,i,j):
    try:
        workbook = xlrd.open_workbook(path)
        worksheet = workbook.sheet_by_name(sheet)
        data = worksheet.cell_value(i,j)
    except Exception as msg:
        errorMsg = msg
    return data

#获取excel的行数
def getTotalRows(path,sheet):
    try:
        workbook=xlrd.open_workbook(path)
        worksheet=workbook.sheet_by_name(sheet)
        rows = worksheet.nrows
        print(rows)
    except Exception as msg:
        errorMsg = msg
    return rows


#根据列名得到的对应得列数
def getColumn(path,sheet,columnName):
    try:
        workbook=xlrd.open_workbook(path)
        worksheet=workbook.sheet_by_name(sheet)
        for i in range(worksheet.ncols):
            if worksheet.cell(0,i) == columnName:
                column = i
                break
    except Exception as msg:
        errorMsg = msg
    return column


#向excel中存数据
def inputExcel(path,sheet,i,j,data):
    try:
        wb=load_workbook(path)
        wb1=wb.active
        wb1.cell(i,j,data)
        wb.save(path)
    except Exception as msg:
        errorMsg=msg


#json转换为字典
def jsonToDic(json1):
    result = json.loads(json1,strict=False)
    return result


#字典转换为json
def dicToJson(dic):
    result=json.dumps(dic,indent=2,sort_keys=True)
    return result




